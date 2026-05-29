"""
app/services/export_service.py — Экспорт результатов распределения в Excel

Структура Excel-отчёта (до 16 листов):
──────────────────────────────────────────────────────────────────────────────

  Результаты распределения:
    «Распределение»                     — главная таблица (один ряд = одна потребность)
    «Движение склада»                   — FIFO-расход по партиям
    «Остатки складов (итог)» — состояние складов после расчёта
    «Возможное перемещение»             — склады других филиалов для анализа (не факт)
    «Шаблон подтверждения»  — шаблон для руководителя, + колонка «Кол-во согласовано»
    «Одобренные перемещения»            — реально использованные перемещения (Слой 3в)
    «Обеспеченность»                    — сводка по работам: покрыто / дефицит / %
    «Распределение «в пути»»            — распределённые поставки
    «Не распред. в пути»                — нераспределённые остатки поставок
    «Дефицит»                           — материалы к закупу (все необходимые поля)

  Исходные данные (для аудита и сверки):
    «Исх. Перечень работ»     — все работы с датами, статусами, кол-вом материалов
    «Исх. Потребности»        — все потребности до распределения
    «Исх. Остатки складов»    — начальные остатки до распределения
    «Исх. Поставки»           — все строки поставок
    «Исх. Списания»           — фактические списания (Слой 1)
    «Исх. Выдано не списано»  — выданные, но не списанные материалы (Слой 2)

  Форматирование:
    - Заголовок: жирный текст, светло-синий фон (#D9E1F2), высота 32
    - Числа: числовой формат Excel (#,##0.0000 / #,##0.00 / %)
    - Даты:  строка ДД.ММ.ГГГГ
    - Автоширина колонок [10, 55]
    - Закреплена первая строка (freeze_panes)
"""

from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Optional

import pandas as pd
from sqlalchemy import text

from app.core.config import settings
from app.core.logging_config import get_logger
from app.db.database import get_session

logger = get_logger(__name__)

# Форматы чисел для Excel (openpyxl)
_FMT_QTY   = '#,##0.0000'   # количества (4 знака)
_FMT_PRICE = '#,##0.00'     # цены и суммы (2 знака)
_FMT_PCT   = '0.0'          # проценты (1 знак)
_FMT_INT   = '0'            # целые числа

# Ключи для раздельного экспорта отдельных листов
SHEET_KEYS = {
    "distribution":  "Распределение",
    "movements":     "Движение склада",
    "balances":      "Остатки складов (итог)",
    "possible":      "Возможное перемещение",
    "transfers_tmpl":"Шаблон подтверждения",
    "approved":      "Одобренные перемещения",
    "coverage":      "Обеспеченность",
    "in_transit":    "Распределение «в пути»",
    "in_transit_free":"Не распред. в пути",
    "deficit":       "Дефицит",
    "src_works":     "Исх. Перечень работ",
    "src_req":       "Исх. Потребности",
    "src_stock":     "Исх. Остатки складов",
    "src_supplies":  "Исх. Поставки",
    "src_writeoffs": "Исх. Списания",
    "src_issued":    "Исх. Выдано не списано",
}


# =============================================================================
# Вспомогательные функции форматирования
# =============================================================================

def _fmt_date(val: object) -> str:
    """
    Форматировать дату в формате ДД.ММ.ГГГГ.

    Принимает: date, datetime, pandas Timestamp, строку.
    Возвращает строку "28.05.2026" или "" если пусто.
    """
    if val is None:
        return ""
    try:
        if pd.isna(val):  # type: ignore[arg-type]
            return ""
    except (TypeError, ValueError):
        pass
    if hasattr(val, "strftime"):
        return val.strftime("%d.%m.%Y")  # type: ignore[union-attr]
    s = str(val).strip()
    if not s or s in ("None", "NaT", "nan"):
        return ""
    for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
        try:
            from datetime import datetime as dt
            return dt.strptime(s, fmt).strftime("%d.%m.%Y")
        except ValueError:
            continue
    return s



def export_single_sheet(
    session_id: str,
    sheet_key: str,
    output_dir: Optional[str] = None,
) -> Path:
    """
    Сформировать Excel-файл с одним листом отчёта.

    Args:
        session_id: ID сессии распределения
        sheet_key:  Ключ листа из SHEET_KEYS (например "deficit", "distribution")
        output_dir: Папка для сохранения (по умолчанию из settings.export_dir)

    Returns:
        Path — путь к созданному .xlsx файлу

    Raises:
        ValueError: если sheet_key не найден в SHEET_KEYS
    """
    if sheet_key not in SHEET_KEYS:
        raise ValueError(
            f"Неизвестный ключ листа: {sheet_key!r}. Допустимые: {list(SHEET_KEYS)}"
        )

    sheet_name = SHEET_KEYS[sheet_key]
    out_dir = Path(output_dir or settings.export_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = out_dir / f"maps_{sheet_key}_{session_id}_{timestamp}.xlsx"

    logger.info("Формирование листа '%s' | сессия: %s", sheet_name, session_id)

    _sheet_builders = {
        "distribution":   lambda s: _get_wide_allocation_df(s, session_id),
        "movements":      lambda s: _get_movement_df(s, session_id),
        "balances":       lambda s: _get_warehouse_balances_df(s, session_id),
        "possible":       lambda s: _get_possible_movements_df(s, session_id),
        "transfers_tmpl": lambda s: _get_transfer_confirmation_df(s, session_id),
        "approved":       lambda s: _get_approved_transfers_df(s, session_id),
        "coverage":       lambda s: _get_coverage_by_work_df(s, session_id),
        "in_transit":     lambda s: _get_in_transit_allocated_df(s, session_id),
        "in_transit_free":lambda s: _get_in_transit_unallocated_df(s),
        "deficit":        lambda s: _get_deficit_df(s, session_id),
        "src_works":      lambda s: _get_source_works_df(s),
        "src_req":        lambda s: _get_source_requirements_df(s),
        "src_stock":      lambda s: _get_source_stock_df(s),
        "src_supplies":   lambda s: _get_source_supplies_df(s),
        "src_writeoffs":  lambda s: _get_source_writeoffs_df(s),
        "src_issued":     lambda s: _get_source_issued_df(s),
    }

    with get_session() as session:
        df = _sheet_builders[sheet_key](session)

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        _write_sheet(writer, df, sheet_name)

    logger.info("Лист '%s' сохранён: %s", sheet_name, file_path)
    return file_path


def export_allocation_results(
    session_id: str,
    output_dir: Optional[str] = None,
) -> Path:
    """
    Сформировать Excel-отчёт по результатам сессии распределения.

    Args:
        session_id: ID сессии распределения (например "20260523_143055_a1b2c")
        output_dir: Папка для сохранения. По умолчанию — из settings.export_dir.

    Returns:
        Path — путь к созданному .xlsx файлу
    """
    out_dir = Path(output_dir or settings.export_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = out_dir / f"maps_results_{session_id}_{timestamp}.xlsx"

    logger.info("Формирование Excel-отчёта | сессия: %s", session_id)

    with get_session() as session:
        df_main         = _get_wide_allocation_df(session, session_id)
        df_movement     = _get_movement_df(session, session_id)
        df_balances     = _get_warehouse_balances_df(session, session_id)
        df_possible     = _get_possible_movements_df(session, session_id)
        df_approved     = _get_approved_transfers_df(session, session_id)
        df_coverage     = _get_coverage_by_work_df(session, session_id)
        df_transit      = _get_in_transit_allocated_df(session, session_id)
        df_transit_free = _get_in_transit_unallocated_df(session, session_id)
        df_deficit      = _get_deficit_df(session, session_id)
        df_tmpl_transf  = _get_transfer_confirmation_df(session, session_id)

        df_src_works    = _get_source_works_df(session)
        df_src_req      = _get_source_requirements_df(session)
        df_src_stock    = _get_source_stock_df(session)
        df_src_supplies = _get_source_supplies_df(session)
        df_src_writeoff = _get_source_writeoffs_df(session)
        df_src_issued   = _get_source_issued_df(session)

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        # Результаты распределения
        _write_sheet(writer, df_main,         "Распределение")
        _write_sheet(writer, df_movement,     "Движение склада")
        _write_sheet(writer, df_balances,     "Остатки складов (итог)")
        _write_sheet(writer, df_possible,     "Возможное перемещение")
        _write_sheet(writer, df_tmpl_transf,  "Шаблон подтверждения")
        if not df_approved.empty:
            _write_sheet(writer, df_approved, "Одобренные перемещения")
        _write_sheet(writer, df_coverage,     "Обеспеченность")
        _write_sheet(writer, df_transit,      "Распределение «в пути»")
        _write_sheet(writer, df_transit_free, "Не распред. в пути")
        _write_sheet(writer, df_deficit,      "Дефицит")

        # Исходные данные (для сверки и аудита)
        _write_sheet(writer, df_src_works,    "Исх. Перечень работ")
        _write_sheet(writer, df_src_req,      "Исх. Потребности")
        _write_sheet(writer, df_src_stock,    "Исх. Остатки складов")
        _write_sheet(writer, df_src_supplies, "Исх. Поставки")
        _write_sheet(writer, df_src_writeoff, "Исх. Списания")
        _write_sheet(writer, df_src_issued,   "Исх. Выдано не списано")

    logger.info(
        "Отчёт создан: %s | строк в главной таблице: %d",
        file_path, len(df_main),
    )
    return file_path


# =============================================================================
# Лист 1: Главная широкая таблица (все слои в одной строке)
# =============================================================================

def _get_wide_allocation_df(session, session_id: str) -> pd.DataFrame:
    """
    Сформировать главную широкую таблицу распределения.

    Одна строка = одна потребность (work + material).
    Колонки: данные работы → материал → потребность → слои 1-4 → к закупу → %.
    Числа передаются как float — форматирование делает Excel.
    Полные названия слоёв начиная с колонки «Системный номер» (N по счёту).
    """
    reqs_sql = text("""
        SELECT
            r.id                            AS requirement_id,
            r.potrebnost,
            r.prognosnaya_tsena,
            r.raspredeleno,
            r.deficit,
            w.is_emergency,
            w.prioritet,
            w.kod_raboty,
            w.filial                        AS filial_raboty,
            w.zavod                         AS zavod_raboty,
            w.data_nachala,
            w.data_okonchaniya,
            m.sys_nomer,
            m.naimenovanie,
            m.ed_izm
        FROM requirements r
        JOIN works     w ON r.work_id     = w.id
        JOIN materials m ON r.material_id = m.id
        ORDER BY
            w.is_emergency DESC,
            w.data_nachala ASC NULLS LAST,
            CASE WHEN NOT w.is_emergency THEN w.prioritet ELSE 0 END ASC,
            w.filial ASC NULLS LAST,
            w.kod_raboty ASC,
            m.sys_nomer ASC
    """)
    reqs_rows = session.execute(reqs_sql).fetchall()
    if not reqs_rows:
        return pd.DataFrame()

    df_reqs = pd.DataFrame(reqs_rows, columns=list(reqs_rows[0]._fields))
    for _col in ["potrebnost", "prognosnaya_tsena", "raspredeleno", "deficit"]:
        if _col in df_reqs.columns:
            df_reqs[_col] = df_reqs[_col].astype(float)

    alloc_sql = text("""
        SELECT
            ar.requirement_id,
            ar.istochnik,
            ar.kolichestvo,
            ar.srednyaya_stoimost,
            ar.summa,
            s.dogovor,
            s.postavshchik
        FROM allocation_results ar
        LEFT JOIN supply_lines sl ON ar.supply_line_id = sl.id
        LEFT JOIN supplies     s  ON sl.supply_id = s.id
        WHERE ar.session_id = :sid
          AND ar.is_possible = FALSE
    """)
    alloc_rows = session.execute(alloc_sql, {"sid": session_id}).fetchall()
    df_alloc = pd.DataFrame(alloc_rows, columns=list(alloc_rows[0]._fields)) if alloc_rows else pd.DataFrame(
        columns=["requirement_id", "istochnik", "kolichestvo", "srednyaya_stoimost", "summa", "dogovor", "postavshchik"]
    )
    for _col in ["kolichestvo", "srednyaya_stoimost", "summa"]:
        if _col in df_alloc.columns:
            df_alloc[_col] = df_alloc[_col].astype(float)

    def _agg_layer(istochnik: str) -> pd.DataFrame:
        df = df_alloc[df_alloc["istochnik"] == istochnik].copy()
        if df.empty:
            return pd.DataFrame(columns=["requirement_id", "qty", "summa"])
        agg = df.groupby("requirement_id").agg(
            qty=("kolichestvo", "sum"),
            summa=("summa", "sum"),
        ).reset_index()
        agg["avg_price"] = agg.apply(
            lambda row: round(float(row["summa"]) / float(row["qty"]), 2)
            if row["qty"] > 0 else 0.0,
            axis=1,
        )
        return agg

    df_sp = _agg_layer("spisanie")
    df_vd = _agg_layer("vydano")
    df_sk = _agg_layer("sklad")
    df_od = _agg_layer("odobren_perenos")

    df_pv_raw = df_alloc[df_alloc["istochnik"] == "postavka"].copy()
    if not df_pv_raw.empty:
        df_pv = df_pv_raw.groupby("requirement_id").agg(
            qty=("kolichestvo", "sum"),
            summa=("summa", "sum"),
            dogovora=("dogovor", lambda x: ", ".join(filter(None, x.dropna().unique().tolist()))),
            postavshchiki=("postavshchik", lambda x: ", ".join(filter(None, x.dropna().unique().tolist()))),
        ).reset_index()
        df_pv["avg_price"] = df_pv.apply(
            lambda row: round(float(row["summa"]) / float(row["qty"]), 2)
            if row["qty"] > 0 else 0.0,
            axis=1,
        )
    else:
        df_pv = pd.DataFrame(
            columns=["requirement_id", "qty", "summa", "avg_price", "dogovora", "postavshchiki"]
        )

    df = df_reqs.copy()
    df["prognosnaya_stoimost"] = df["potrebnost"].astype(float) * df["prognosnaya_tsena"].astype(float)

    def _merge_layer(df_main: pd.DataFrame, df_layer: pd.DataFrame, prefix: str) -> pd.DataFrame:
        if df_layer.empty:
            df_main[f"{prefix}_qty"] = 0.0
            df_main[f"{prefix}_price"] = 0.0
            df_main[f"{prefix}_summa"] = 0.0
            return df_main
        df_layer = df_layer.rename(columns={
            "qty": f"{prefix}_qty",
            "avg_price": f"{prefix}_price",
            "summa": f"{prefix}_summa",
        })
        df_merged = df_main.merge(
            df_layer[["requirement_id", f"{prefix}_qty", f"{prefix}_price", f"{prefix}_summa"]],
            on="requirement_id",
            how="left",
        )
        for col in [f"{prefix}_qty", f"{prefix}_price", f"{prefix}_summa"]:
            df_merged[col] = df_merged[col].fillna(0.0)
        return df_merged

    df = _merge_layer(df, df_sp, "sp")
    df = _merge_layer(df, df_vd, "vd")
    df = _merge_layer(df, df_sk, "sk")
    df = _merge_layer(df, df_od, "od")

    if not df_pv.empty:
        df_pv_merged = df_pv.rename(columns={
            "qty": "pv_qty",
            "avg_price": "pv_price",
            "summa": "pv_summa",
        })
        df = df.merge(
            df_pv_merged[["requirement_id", "pv_qty", "pv_price", "pv_summa", "dogovora", "postavshchiki"]],
            on="requirement_id",
            how="left",
        )
        for col in ["pv_qty", "pv_price", "pv_summa"]:
            df[col] = df[col].fillna(0.0)
        df["dogovora"] = df["dogovora"].fillna("")
        df["postavshchiki"] = df["postavshchiki"].fillna("")
    else:
        df["pv_qty"] = 0.0
        df["pv_price"] = 0.0
        df["pv_summa"] = 0.0
        df["dogovora"] = ""
        df["postavshchiki"] = ""

    df["covered_summa"] = (
        df["sp_summa"] + df["vd_summa"] + df["sk_summa"]
        + df["od_summa"] + df["pv_summa"]
    )

    df["zakup_qty"]   = df["deficit"].astype(float)
    df["zakup_price"] = df["prognosnaya_tsena"].astype(float)
    df["zakup_summa"] = df["zakup_qty"] * df["zakup_price"]

    df["coverage_pct"] = df.apply(
        lambda row: (
            round(row["covered_summa"] / row["prognosnaya_stoimost"] * 100, 1)
            if row["prognosnaya_stoimost"] > 0
            else (
                round(row["raspredeleno"] / row["potrebnost"] * 100, 1)
                if float(row["potrebnost"]) > 0 else 0.0
            )
        ),
        axis=1,
    )

    # Числа — сырые float; Excel-форматирование применяет _write_sheet.
    # Полные названия слоёв начиная со столбца «Системный номер» (колонка N).
    result = pd.DataFrame({
        # Признак аварийности
        "Аварийная":    df["is_emergency"].apply(lambda x: "Да" if x else ""),

        # Данные работы
        "Приоритет":        df["prioritet"],
        "Код работы":       df["kod_raboty"],
        "Филиал":           df["filial_raboty"],
        "Завод":            df["zavod_raboty"],
        "Дата начала":      df["data_nachala"].apply(_fmt_date),
        "Дата окончания":   df["data_okonchaniya"].apply(_fmt_date),

        # Данные материала (колонка H — с этого места нет сокращений)
        "Системный номер":  df["sys_nomer"],
        "Наименование":     df["naimenovanie"],
        "Ед.изм":           df["ed_izm"],

        # Потребность
        "Потребность":          df["potrebnost"].round(4),
        "Прогнозная цена":      df["prognosnaya_tsena"].round(2),
        "Прогнозная стоимость": df["prognosnaya_stoimost"].round(2),

        # Слой 1: Списание (полные названия)
        "Списание: Количество": df["sp_qty"].round(4),
        "Списание: Цена":       df["sp_price"].round(2),
        "Списание: Сумма":      df["sp_summa"].round(2),

        # Слой 2: Выдано не списано
        "Выдано: Количество":   df["vd_qty"].round(4),
        "Выдано: Цена":         df["vd_price"].round(2),
        "Выдано: Сумма":        df["vd_summa"].round(2),

        # Слой 3: Склад своего завода/филиала
        "Склад: Количество":    df["sk_qty"].round(4),
        "Склад: Средняя цена":  df["sk_price"].round(2),
        "Склад: Сумма":         df["sk_summa"].round(2),

        # Слой 3в: Одобренные межфилиальные перемещения
        "Одобр. перемещение: Количество":   df["od_qty"].round(4),
        "Одобр. перемещение: Средняя цена": df["od_price"].round(2),
        "Одобр. перемещение: Сумма":        df["od_summa"].round(2),

        # Слой 4: Поставки
        "Поставка: Договор(а)":   df["dogovora"],
        "Поставка: Поставщик(и)": df["postavshchiki"],
        "Поставка: Количество":   df["pv_qty"].round(4),
        "Поставка: Средняя цена": df["pv_price"].round(2),
        "Поставка: Сумма":        df["pv_summa"].round(2),

        # К закупу
        "К закупу: Количество": df["zakup_qty"].round(4),
        "К закупу: Цена":       df["zakup_price"].round(2),
        "К закупу: Сумма":      df["zakup_summa"].round(2),

        # Итоговый показатель
        "Обеспечённость %": df["coverage_pct"].round(1),
    })

    return result


# =============================================================================
# Лист 2: Движение склада (детальное, по партиям)
# =============================================================================

def _get_movement_df(session, session_id: str) -> pd.DataFrame:
    """
    Детальное движение по складам — каждое списание с партии на работу.

    Показывает цепочку:
        Склад W-001 → Партия 0045 → Работа WO-123 → Материал 00010012345 → 50 м → Остаток: 150 м

    Добавлено поле «Дата начала работы» — дата начала работы, на которую
    были распределены остатки.
    """
    sql = text("""
        SELECT
            wh.kod_sklada                           AS "Склад",
            wh.filial                               AS "Филиал склада",
            wh.zavod                                AS "Завод склада",
            m.sys_nomer                             AS "Системный номер",
            m.naimenovanie                          AS "Наименование материала",
            m.ed_izm                                AS "Ед.изм",
            sb.nomer_partii                         AS "Номер партии",
            sb.data_postupleniya                    AS "Дата поступления партии",
            CAST(sb.kolichestvo AS NUMERIC(18,4))   AS "Нач. кол-во партии",
            CAST(sb.stoimost_za_ed AS NUMERIC(18,2)) AS "Цена за ед.",
            w.kod_raboty                            AS "Работа-получатель",
            w.filial                                AS "Филиал работы",
            w.zavod                                 AS "Завод работы",
            w.prioritet                             AS "Приоритет работы",
            w.is_emergency                          AS "Аварийная работа",
            w.data_nachala                          AS "Дата начала работы",
            CAST(ABS(sm.izmenenie) AS NUMERIC(18,4)) AS "Списано",
            CAST(sm.ostatok AS NUMERIC(18,4))        AS "Остаток партии после",
            sm.data_dvizheniya                      AS "Дата движения"
        FROM stock_movements sm
        JOIN warehouses  wh ON sm.warehouse_id = wh.id
        JOIN materials    m ON sm.material_id  = m.id
        LEFT JOIN stock_batches sb ON sm.batch_id  = sb.id
        LEFT JOIN works          w ON sm.work_id   = w.id
        WHERE sm.session_id = :sid
        ORDER BY
            wh.kod_sklada,
            m.sys_nomer,
            sb.data_postupleniya,
            sm.id
    """)
    rows = session.execute(sql, {"sid": session_id}).fetchall()
    df = _to_df(rows)
    if df.empty:
        return df
    for col in ["Дата поступления партии", "Дата движения", "Дата начала работы"]:
        if col in df.columns:
            df[col] = df[col].apply(_fmt_date)
    return df


# =============================================================================
# Лист 3: Остатки складов (итог)
# =============================================================================

def _get_warehouse_balances_df(session, _session_id: str = "") -> pd.DataFrame:
    """
    Текущие остатки по складам ПОСЛЕ распределения.

    Показывает для каждой партии:
        • Начальный остаток (kolichestvo — не менялся при импорте)
        • Распределено в этой сессии = kolichestvo - dostupno
        • Текущий остаток (dostupno — обновлён алгоритмом)
        • Сумма остатка = dostupno × stoimost_za_ed
    """
    sql = text("""
        SELECT
            wh.kod_sklada                           AS "Склад",
            wh.filial                               AS "Филиал склада",
            wh.zavod                                AS "Завод склада",
            m.sys_nomer                             AS "Системный номер",
            m.naimenovanie                          AS "Наименование материала",
            m.ed_izm                                AS "Ед.изм",
            m.gruppa                                AS "Группа",
            sb.nomer_partii                         AS "Номер партии",
            sb.data_postupleniya                    AS "Дата поступления",
            CAST(sb.kolichestvo AS NUMERIC(18,4))   AS "Нач. остаток",
            CAST(sb.kolichestvo - sb.dostupno AS NUMERIC(18,4)) AS "Распределено",
            CAST(sb.dostupno AS NUMERIC(18,4))      AS "Текущий остаток",
            CAST(sb.stoimost_za_ed AS NUMERIC(18,2)) AS "Цена за ед.",
            CAST(sb.dostupno * sb.stoimost_za_ed AS NUMERIC(18,2)) AS "Сумма остатка"
        FROM stock_batches sb
        JOIN warehouses wh ON sb.warehouse_id = wh.id
        JOIN materials   m ON sb.material_id  = m.id
        ORDER BY
            wh.kod_sklada,
            m.sys_nomer,
            sb.data_postupleniya
    """)
    rows = session.execute(sql).fetchall()
    df = _to_df(rows)
    if df.empty:
        return df
    if "Дата поступления" in df.columns:
        df["Дата поступления"] = df["Дата поступления"].apply(_fmt_date)
    return df


# =============================================================================
# Лист 4: Возможное перемещение из других филиалов
# =============================================================================

def _get_possible_movements_df(session, session_id: str) -> pd.DataFrame:
    """
    Возможное покрытие дефицита из складов других филиалов.

    is_possible=TRUE — не фактическое распределение, только для анализа.
    """
    sql = text("""
        SELECT
            w.prioritet                             AS "Приоритет",
            w.is_emergency                          AS "Аварийная",
            w.kod_raboty                            AS "Код работы",
            w.filial                                AS "Филиал работы",
            w.zavod                                 AS "Завод работы",
            w.data_nachala                          AS "Дата начала",
            w.data_okonchaniya                      AS "Нужно к дате",
            m.sys_nomer                             AS "Системный номер",
            m.naimenovanie                          AS "Наименование материала",
            m.ed_izm                                AS "Ед.изм",
            wh.kod_sklada                           AS "Склад-источник",
            wh.filial                               AS "Филиал склада",
            wh.zavod                                AS "Завод склада",
            CAST(ar.kolichestvo AS NUMERIC(18,4))        AS "Возможное кол-во",
            CAST(ar.srednyaya_stoimost AS NUMERIC(18,2)) AS "Средняя цена",
            CAST(ar.summa AS NUMERIC(18,2))              AS "Оценочная сумма",
            'Требует согласования'                  AS "Статус"
        FROM allocation_results ar
        JOIN works     w  ON ar.work_id     = w.id
        JOIN materials m  ON ar.material_id = m.id
        LEFT JOIN warehouses wh ON ar.warehouse_id = wh.id
        WHERE ar.session_id = :sid
          AND ar.is_possible = TRUE
          AND ar.istochnik = 'vozmozhnoe_sklad'
        ORDER BY
            w.is_emergency DESC,
            w.data_nachala ASC NULLS LAST,
            CASE WHEN NOT w.is_emergency THEN w.prioritet ELSE 0 END ASC,
            w.kod_raboty,
            m.sys_nomer
    """)
    rows = session.execute(sql, {"sid": session_id}).fetchall()
    df = _to_df(rows)
    if df.empty:
        return df
    for col in ["Дата начала", "Нужно к дате"]:
        if col in df.columns:
            df[col] = df[col].apply(_fmt_date)
    return df


# =============================================================================
# Лист «Одобренные перемещения» (результат Layer 3в)
# =============================================================================

def _get_approved_transfers_df(session, session_id: str) -> pd.DataFrame:
    """
    Список согласованных перемещений, которые были использованы в распределении.
    """
    sql = text("""
        SELECT
            w.prioritet                                  AS "Приоритет",
            w.is_emergency                               AS "Аварийная",
            w.kod_raboty                                 AS "Код работы",
            w.filial                                     AS "Филиал работы",
            w.zavod                                      AS "Завод работы",
            w.data_nachala                               AS "Дата начала",
            w.data_okonchaniya                           AS "Нужно к дате",
            m.sys_nomer                                  AS "Системный номер",
            m.naimenovanie                               AS "Наименование материала",
            m.ed_izm                                     AS "Ед.изм",
            wh.kod_sklada                                AS "Склад-источник",
            wh.filial                                    AS "Филиал склада",
            wh.zavod                                     AS "Завод склада",
            CAST(ar.kolichestvo AS NUMERIC(18,4))        AS "Взято кол-во",
            CAST(ar.srednyaya_stoimost AS NUMERIC(18,2)) AS "Средняя цена",
            CAST(ar.summa AS NUMERIC(18,2))              AS "Сумма",
            'Одобрено руководством'                      AS "Статус"
        FROM allocation_results ar
        JOIN works     w  ON ar.work_id     = w.id
        JOIN materials m  ON ar.material_id = m.id
        LEFT JOIN warehouses wh ON ar.warehouse_id = wh.id
        WHERE ar.session_id = :sid
          AND ar.istochnik  = 'odobren_perenos'
        ORDER BY
            w.is_emergency DESC,
            w.data_nachala ASC NULLS LAST,
            w.kod_raboty,
            m.sys_nomer
    """)
    rows = session.execute(sql, {"sid": session_id}).fetchall()
    df = _to_df(rows)
    if df.empty:
        return df
    for col in ["Дата начала", "Нужно к дате"]:
        if col in df.columns:
            df[col] = df[col].apply(_fmt_date)
    return df


# =============================================================================
# Лист 5: Обеспеченность по работам (сводная таблица)
# =============================================================================

def _get_coverage_by_work_df(session, session_id: str) -> pd.DataFrame:
    """
    Сводная таблица обеспеченности по работам.

    Одна строка = одна работа. Показывает: стоимость потребности,
    стоимость обеспеченного, стоимость дефицита, % обеспеченности.
    """
    sql = text("""
        SELECT
            w.kod_raboty,
            w.nazvanie,
            w.filial,
            w.zavod,
            w.data_nachala,
            w.data_okonchaniya,
            w.is_emergency,
            COUNT(r.id)                                         AS materials_count,
            COALESCE(SUM(r.potrebnost * r.prognosnaya_tsena), 0) AS stoimost_potrebnosti,
            COALESCE((
                SELECT SUM(ar.summa)
                FROM allocation_results ar
                WHERE ar.work_id = w.id
                  AND ar.session_id = :sid
                  AND ar.is_possible = FALSE
            ), 0)                                               AS stoimost_obespecheno,
            COALESCE((
                SELECT SUM(dr.estimated_cost)
                FROM deficit_records dr
                WHERE dr.work_id = w.id
                  AND dr.session_id = :sid
            ), 0)                                               AS stoimost_deficita
        FROM works w
        JOIN requirements r ON r.work_id = w.id
        GROUP BY w.id, w.kod_raboty, w.nazvanie, w.filial, w.zavod,
                 w.data_nachala, w.data_okonchaniya, w.is_emergency
        ORDER BY w.data_nachala ASC NULLS LAST, w.kod_raboty
    """)

    rows = session.execute(sql, {"sid": session_id}).fetchall()
    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows, columns=list(rows[0]._fields))
    for col in ["stoimost_potrebnosti", "stoimost_obespecheno", "stoimost_deficita"]:
        df[col] = df[col].astype(float)

    df["coverage_pct"] = df.apply(
        lambda row: (
            round(float(row["stoimost_obespecheno"]) / float(row["stoimost_potrebnosti"]) * 100, 1)
            if float(row["stoimost_potrebnosti"]) > 0
            else 0.0
        ),
        axis=1,
    )

    result = pd.DataFrame({
        "Код работы":           df["kod_raboty"],
        "Наименование работы":  df["nazvanie"].fillna(""),
        "Филиал":               df["filial"].fillna(""),
        "Завод":                df["zavod"].fillna(""),
        "Дата начала":          df["data_nachala"].apply(_fmt_date),
        "Дата окончания":       df["data_okonchaniya"].apply(_fmt_date),
        "Аварийная":            df["is_emergency"].apply(lambda x: "Да" if x else "Нет"),
        "Материалов":           df["materials_count"].astype(int),
        "Стоимость потребности": df["stoimost_potrebnosti"].round(2),
        "Стоимость обеспечено":  df["stoimost_obespecheno"].round(2),
        "Стоимость дефицита":    df["stoimost_deficita"].round(2),
        "Обеспеченность %":      df["coverage_pct"].round(1),
    })

    return result


# =============================================================================
# Лист «Распределение «в пути»» — распределённые поставки
# =============================================================================

def _get_in_transit_allocated_df(session, session_id: str) -> pd.DataFrame:
    """
    Строки поставок («в пути»), распределённые на потребности в данной сессии.

    Колонка «Наименование работы» убрана — дублирует «Обеспеченность».
    """
    sql = text("""
        SELECT
            s.dogovor                           AS "Договор",
            s.postavshchik                      AS "Поставщик",
            s.data_postavki                     AS "Дата поставки",
            s.filial                            AS "Филиал поставки",
            s.zavod                             AS "Завод поставки",
            w.kod_raboty                        AS "Код работы",
            m.sys_nomer                         AS "Системный номер",
            m.naimenovanie                      AS "Наименование материала",
            m.ed_izm                            AS "Ед.изм",
            CAST(ar.kolichestvo AS NUMERIC(18,4))      AS "Кол-во распределено",
            CAST(ar.srednyaya_stoimost AS NUMERIC(18,2)) AS "Стоимость за ед",
            CAST(ar.summa AS NUMERIC(18,2))             AS "Сумма"
        FROM allocation_results ar
        JOIN supply_lines sl  ON sl.id  = ar.supply_line_id
        JOIN supplies s       ON s.id   = sl.supply_id
        JOIN requirements r   ON r.id   = ar.requirement_id
        JOIN works w          ON w.id   = r.work_id
        JOIN materials m      ON m.id   = ar.material_id
        WHERE ar.session_id   = :sid
          AND ar.istochnik    = 'postavka'
        ORDER BY s.data_postavki ASC NULLS LAST,
                 s.dogovor,
                 w.kod_raboty,
                 m.sys_nomer
    """)
    rows = session.execute(sql, {"sid": session_id}).fetchall()
    if not rows:
        return pd.DataFrame(columns=[
            "Договор", "Поставщик", "Дата поставки", "Филиал поставки", "Завод поставки",
            "Код работы", "Системный номер", "Наименование материала",
            "Ед.изм", "Кол-во распределено", "Стоимость за ед", "Сумма",
        ])
    df = _to_df(rows)
    if "Дата поставки" in df.columns:
        df["Дата поставки"] = df["Дата поставки"].apply(_fmt_date)
    return df


# =============================================================================
# Лист 7: Не распределено в пути
# =============================================================================

def _get_in_transit_unallocated_df(session, _session_id: str = "") -> pd.DataFrame:
    """
    Строки поставок с остатком dostupno > 0 после распределения.
    """
    sql = text("""
        SELECT
            s.dogovor                           AS "Договор",
            s.postavshchik                      AS "Поставщик",
            s.data_postavki                     AS "Дата поставки",
            s.filial                            AS "Филиал",
            s.zavod                             AS "Завод",
            m.sys_nomer                         AS "Системный номер",
            m.naimenovanie                      AS "Наименование материала",
            m.ed_izm                            AS "Ед.изм",
            CAST(sl.kolichestvo AS NUMERIC(18,4))              AS "Кол-во всего",
            CAST(sl.kolichestvo - sl.dostupno AS NUMERIC(18,4)) AS "Распределено",
            CAST(sl.dostupno AS NUMERIC(18,4))                 AS "Остаток",
            CAST(sl.stoimost_za_ed AS NUMERIC(18,2))           AS "Стоимость за ед",
            CAST(sl.dostupno * sl.stoimost_za_ed AS NUMERIC(18,2)) AS "Сумма остатка"
        FROM supply_lines sl
        JOIN supplies s   ON s.id  = sl.supply_id
        JOIN materials m  ON m.id  = sl.material_id
        WHERE sl.dostupno > 0
        ORDER BY s.data_postavki ASC NULLS LAST,
                 s.dogovor,
                 m.sys_nomer
    """)
    rows = session.execute(sql).fetchall()
    if not rows:
        return pd.DataFrame(columns=[
            "Договор", "Поставщик", "Дата поставки", "Филиал", "Завод",
            "Системный номер", "Наименование материала", "Ед.изм",
            "Кол-во всего", "Распределено", "Остаток", "Стоимость за ед", "Сумма остатка",
        ])
    df = _to_df(rows)
    if "Дата поставки" in df.columns:
        df["Дата поставки"] = df["Дата поставки"].apply(_fmt_date)
    return df


# =============================================================================
# Листы с исходными данными (для сверки и аудита)
# =============================================================================

def _get_source_requirements_df(session) -> pd.DataFrame:
    """
    Лист «Исх. Потребности» — все потребности до распределения.

    Колонка «Наименование работы» убрана (используй лист «Обеспеченность»
    для просмотра наименований работ).
    """
    sql = text("""
        SELECT
            w.is_emergency                               AS "Аварийная",
            w.prioritet                                  AS "Приоритет",
            w.kod_raboty                                 AS "Код работы",
            w.filial                                     AS "Филиал",
            w.zavod                                      AS "Завод",
            w.data_nachala                               AS "Дата начала",
            w.data_okonchaniya                           AS "Дата окончания",
            w.status                                     AS "Статус работы",
            m.sys_nomer                                  AS "Системный номер",
            m.naimenovanie                               AS "Наименование материала",
            m.ed_izm                                     AS "Ед.изм",
            CAST(r.potrebnost AS NUMERIC(18,4))          AS "Потребность",
            CAST(r.prognosnaya_tsena AS NUMERIC(18,2))   AS "Прогнозная цена",
            CAST(r.potrebnost * r.prognosnaya_tsena AS NUMERIC(18,2)) AS "Прогнозная стоимость"
        FROM requirements r
        JOIN works     w ON r.work_id     = w.id
        JOIN materials m ON r.material_id = m.id
        ORDER BY
            w.is_emergency DESC,
            w.data_nachala ASC NULLS LAST,
            w.prioritet ASC,
            w.kod_raboty,
            m.sys_nomer
    """)
    rows = session.execute(sql).fetchall()
    df = _to_df(rows)
    if df.empty:
        return df
    df["Аварийная"] = df["Аварийная"].apply(lambda x: "Да" if x else "")
    for col in ["Дата начала", "Дата окончания"]:
        if col in df.columns:
            df[col] = df[col].apply(_fmt_date)
    return df


def _get_source_stock_df(session) -> pd.DataFrame:
    """
    Лист «Исх. Остатки складов» — начальные остатки до распределения.

    Показывает kolichestvo (полный исходный остаток, не изменяется алгоритмом).
    """
    sql = text("""
        SELECT
            wh.kod_sklada                               AS "Склад",
            wh.filial                                   AS "Филиал склада",
            wh.zavod                                    AS "Завод склада",
            m.sys_nomer                                 AS "Системный номер",
            m.naimenovanie                              AS "Наименование материала",
            m.ed_izm                                    AS "Ед.изм",
            m.gruppa                                    AS "Группа",
            sb.nomer_partii                             AS "Номер партии",
            sb.data_postupleniya                        AS "Дата поступления",
            CAST(sb.kolichestvo AS NUMERIC(18,4))       AS "Кол-во (исходное)",
            CAST(sb.stoimost_za_ed AS NUMERIC(18,2))    AS "Цена за ед.",
            CAST(sb.kolichestvo * sb.stoimost_za_ed AS NUMERIC(18,2)) AS "Сумма"
        FROM stock_batches sb
        JOIN warehouses wh ON sb.warehouse_id = wh.id
        JOIN materials   m ON sb.material_id  = m.id
        ORDER BY
            wh.kod_sklada,
            m.sys_nomer,
            sb.data_postupleniya
    """)
    rows = session.execute(sql).fetchall()
    df = _to_df(rows)
    if df.empty:
        return df
    if "Дата поступления" in df.columns:
        df["Дата поступления"] = df["Дата поступления"].apply(_fmt_date)
    return df


def _get_source_supplies_df(session) -> pd.DataFrame:
    """
    Лист «Исх. Поставки» — все строки поставок, загруженные в систему.
    """
    sql = text("""
        SELECT
            s.dogovor                                   AS "Договор",
            s.postavshchik                              AS "Поставщик",
            s.data_postavki                             AS "Дата поставки",
            s.filial                                    AS "Филиал",
            s.zavod                                     AS "Завод",
            s.status                                    AS "Статус поставки",
            m.sys_nomer                                 AS "Системный номер",
            m.naimenovanie                              AS "Наименование материала",
            m.ed_izm                                    AS "Ед.изм",
            CAST(sl.kolichestvo AS NUMERIC(18,4))       AS "Кол-во",
            CAST(sl.stoimost_za_ed AS NUMERIC(18,2))    AS "Цена за ед.",
            CAST(sl.kolichestvo * sl.stoimost_za_ed AS NUMERIC(18,2)) AS "Сумма"
        FROM supply_lines sl
        JOIN supplies  s ON s.id = sl.supply_id
        JOIN materials m ON m.id = sl.material_id
        ORDER BY
            s.data_postavki ASC NULLS LAST,
            s.dogovor,
            m.sys_nomer
    """)
    rows = session.execute(sql).fetchall()
    df = _to_df(rows)
    if df.empty:
        return df
    if "Дата поставки" in df.columns:
        df["Дата поставки"] = df["Дата поставки"].apply(_fmt_date)
    return df


def _get_source_works_df(session) -> pd.DataFrame:
    """
    Лист «Исх. Перечень работ» — все работы с их свойствами.
    """
    sql = text("""
        SELECT
            CASE WHEN w.is_emergency THEN 'Да' ELSE '' END AS "Аварийная",
            w.prioritet                                     AS "Приоритет",
            w.kod_raboty                                    AS "Код работы",
            w.nazvanie                                      AS "Наименование работы",
            w.filial                                        AS "Филиал",
            w.zavod                                         AS "Завод",
            w.data_nachala                                  AS "Дата начала",
            w.data_okonchaniya                              AS "Дата окончания",
            w.status                                        AS "Статус",
            COUNT(r.id)                                     AS "Кол-во материалов"
        FROM works w
        LEFT JOIN requirements r ON r.work_id = w.id
        GROUP BY w.id, w.is_emergency, w.prioritet, w.kod_raboty, w.nazvanie,
                 w.filial, w.zavod, w.data_nachala, w.data_okonchaniya, w.status
        ORDER BY w.is_emergency DESC, w.data_nachala ASC NULLS LAST, w.prioritet, w.kod_raboty
    """)
    rows = session.execute(sql).fetchall()
    df = _to_df(rows)
    if df.empty:
        return df
    for col in ["Дата начала", "Дата окончания"]:
        if col in df.columns:
            df[col] = df[col].apply(_fmt_date)
    return df


def _get_source_writeoffs_df(session) -> pd.DataFrame:
    """
    Лист «Исх. Списания» — фактические списания до распределения.
    """
    sql = text("""
        SELECT
            w.kod_raboty                                AS "Код работы",
            w.filial                                    AS "Филиал работы",
            m.sys_nomer                                 AS "Системный номер",
            m.naimenovanie                              AS "Наименование материала",
            m.ed_izm                                    AS "Ед.изм",
            CAST(wo.kolichestvo AS NUMERIC(18,4))       AS "Количество",
            CAST(wo.stoimost_za_ed AS NUMERIC(18,2))    AS "Стоимость за ед.",
            CAST(wo.summa AS NUMERIC(18,2))             AS "Сумма",
            wo.nomer_dokumenta                          AS "Номер документа",
            wo.data_spisaniya                           AS "Дата списания"
        FROM writeoffs wo
        JOIN works     w ON wo.work_id     = w.id
        JOIN materials m ON wo.material_id = m.id
        ORDER BY w.kod_raboty, m.sys_nomer
    """)
    rows = session.execute(sql).fetchall()
    df = _to_df(rows)
    if df.empty:
        return df
    if "Дата списания" in df.columns:
        df["Дата списания"] = df["Дата списания"].apply(_fmt_date)
    return df


def _get_source_issued_df(session) -> pd.DataFrame:
    """
    Лист «Исх. Выдано не списано» — выданные, но не списанные материалы.
    """
    sql = text("""
        SELECT
            w.kod_raboty                                AS "Код работы",
            w.filial                                    AS "Филиал работы",
            m.sys_nomer                                 AS "Системный номер",
            m.naimenovanie                              AS "Наименование материала",
            m.ed_izm                                    AS "Ед.изм",
            CAST(iss.kolichestvo AS NUMERIC(18,4))      AS "Количество",
            CAST(iss.stoimost_za_ed AS NUMERIC(18,2))   AS "Стоимость за ед.",
            CAST(iss.summa AS NUMERIC(18,2))            AS "Сумма",
            wh.kod_sklada                               AS "Склад выдачи",
            iss.data_vydachi                            AS "Дата выдачи"
        FROM issued_not_written_off iss
        JOIN works      w  ON iss.work_id     = w.id
        JOIN materials  m  ON iss.material_id = m.id
        LEFT JOIN warehouses wh ON iss.warehouse_id = wh.id
        ORDER BY w.kod_raboty, m.sys_nomer
    """)
    rows = session.execute(sql).fetchall()
    df = _to_df(rows)
    if df.empty:
        return df
    if "Дата выдачи" in df.columns:
        df["Дата выдачи"] = df["Дата выдачи"].apply(_fmt_date)
    return df


def _get_deficit_df(session, session_id: str) -> pd.DataFrame:
    """
    Лист «Дефицит» — материалы, которые не удалось покрыть ни из одного источника.

    Содержит: все необходимые поля для формирования заявки на закупку.
    Одна строка = один материал с дефицитом для одной работы.
    """
    sql = text("""
        SELECT
            w.is_emergency                                          AS "Аварийная",
            w.prioritet                                             AS "Приоритет",
            w.kod_raboty                                            AS "Код работы",
            w.nazvanie                                              AS "Наименование работы",
            w.filial                                                AS "Филиал",
            w.zavod                                                 AS "Завод",
            w.data_nachala                                          AS "Дата начала",
            w.data_okonchaniya                                      AS "Нужно к дате",
            w.status                                                AS "Статус работы",
            m.sys_nomer                                             AS "Системный номер",
            m.naimenovanie                                          AS "Наименование материала",
            m.ed_izm                                                AS "Ед.изм",
            m.gruppa                                                AS "Группа материалов",
            CAST(r.potrebnost AS NUMERIC(18,4))                     AS "Потребность",
            CAST(r.raspredeleno AS NUMERIC(18,4))                   AS "Распределено",
            CAST(dr.deficit_qty AS NUMERIC(18,4))                   AS "Дефицит (кол-во)",
            CAST(r.prognosnaya_tsena AS NUMERIC(18,2))              AS "Прогнозная цена",
            CAST(dr.estimated_cost AS NUMERIC(18,2))                AS "Стоимость дефицита",
            dr.needed_by                                            AS "Нужно к дате (доп.)"
        FROM deficit_records dr
        JOIN works        w  ON dr.work_id        = w.id
        JOIN requirements r  ON dr.requirement_id = r.id
        JOIN materials    m  ON r.material_id      = m.id
        WHERE dr.session_id = :sid
        ORDER BY
            w.is_emergency DESC,
            w.data_nachala ASC NULLS LAST,
            CASE WHEN NOT w.is_emergency THEN w.prioritet ELSE 0 END ASC,
            w.kod_raboty,
            m.sys_nomer
    """)
    rows = session.execute(sql, {"sid": session_id}).fetchall()
    df = _to_df(rows)
    if df.empty:
        return df
    df["Аварийная"] = df["Аварийная"].apply(lambda x: "Да" if x else "")
    for col in ["Дата начала", "Нужно к дате", "Нужно к дате (доп.)"]:
        if col in df.columns:
            df[col] = df[col].apply(_fmt_date)
    return df


def _get_transfer_confirmation_df(session, session_id: str) -> pd.DataFrame:
    """
    Лист «Шаблон подтверждения».

    Шаблон для руководителя: те же данные что и «Возможное перемещение»,
    плюс пустая колонка «Кол-во согласовано» для заполнения вручную.
    Заполненный файл загружается обратно через /api/import/transfers.
    """
    sql = text("""
        SELECT
            w.kod_raboty                                AS "Код работы",
            w.filial                                    AS "Филиал-получатель",
            w.zavod                                     AS "Завод работы",
            w.data_nachala                              AS "Дата начала работы",
            w.data_okonchaniya                          AS "Нужно к дате",
            m.sys_nomer                                 AS "Системный номер",
            m.naimenovanie                              AS "Наименование материала",
            m.ed_izm                                    AS "Ед.изм",
            wh.kod_sklada                               AS "Склад-источник",
            wh.filial                                   AS "Филиал склада-источника",
            CAST(ar.kolichestvo AS NUMERIC(18,4))       AS "Возможное кол-во",
            CAST(ar.srednyaya_stoimost AS NUMERIC(18,2)) AS "Оценочная цена",
            CAST(ar.summa AS NUMERIC(18,2))             AS "Оценочная сумма"
        FROM allocation_results ar
        JOIN works     w  ON ar.work_id     = w.id
        JOIN materials m  ON ar.material_id = m.id
        LEFT JOIN warehouses wh ON ar.warehouse_id = wh.id
        WHERE ar.session_id = :sid
          AND ar.is_possible = TRUE
          AND ar.istochnik = 'vozmozhnoe_sklad'
        ORDER BY
            w.is_emergency DESC,
            w.data_nachala ASC NULLS LAST,
            w.kod_raboty,
            m.sys_nomer
    """)
    rows = session.execute(sql, {"sid": session_id}).fetchall()
    df = _to_df(rows)
    if df.empty:
        return df
    for col in ["Дата начала работы", "Нужно к дате"]:
        if col in df.columns:
            df[col] = df[col].apply(_fmt_date)
    # Добавляем пустую колонку для заполнения — она должна быть последней
    df["Кол-во согласовано"] = ""
    return df


# =============================================================================
# Вспомогательные функции
# =============================================================================

def _to_df(rows) -> pd.DataFrame:
    """
    Преобразовать результат SQL-запроса в DataFrame.

    Decimal-значения (PostgreSQL NUMERIC) автоматически конвертируются в float,
    чтобы pandas хранил их как numeric dtype и Excel получал настоящие числа
    (не строки). Колонки со строками или датами остаются без изменений.
    """
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows, columns=list(rows[0]._fields))
    # Пробуем конвертировать каждую колонку в числовой тип.
    # Если не удаётся (строки, даты, None) — оставляем как есть (errors='ignore').
    for col in df.columns:
        converted = pd.to_numeric(df[col], errors='coerce')
        # Применяем конвертацию только если хотя бы одно значение стало числом
        # и при этом в колонке нет строк с буквами (чтобы не конвертировать коды)
        if not converted.isna().all() and df[col].apply(
            lambda x: isinstance(x, (int, float, Decimal)) if x is not None else True
        ).all():
            df[col] = converted
    return df


_NUM_COL_KEYWORDS = {
    "qty": (
        "кол-во", "количество", "потребность", "нач. остаток", "текущий остаток",
        "нач. кол-во", "распределено", "списано", "остаток", "взято",
        "кол-во всего",
    ),
    "pct": ("%",),
    "int": ("приоритет", "материалов",),
}


def _excel_num_format(col_name: str) -> str:
    """
    Вернуть Excel-формат числа для колонки по её названию.

    Правила:
        «количество / остаток / потребность» → 4 знака после запятой
        «%»                                  → 1 знак
        целые («Приоритет», «Материалов»)    → без дробей
        всё остальное числовое               → 2 знака (цены, суммы)
    """
    lower = col_name.lower()
    for kw in _NUM_COL_KEYWORDS["int"]:
        if kw in lower:
            return _FMT_INT
    for kw in _NUM_COL_KEYWORDS["pct"]:
        if kw in lower:
            return _FMT_PCT
    for kw in _NUM_COL_KEYWORDS["qty"]:
        if kw in lower:
            return _FMT_QTY
    return _FMT_PRICE


def _write_sheet(
    writer: pd.ExcelWriter,
    df: pd.DataFrame,
    sheet_name: str,
) -> None:
    """
    Записать DataFrame в лист Excel с форматированием:
        1. Пустой DataFrame → лист с сообщением «Нет данных»
        2. Заголовок: жирный, светло-синий фон (#D9E1F2), перенос текста
        3. Числовые колонки — Excel number format (#,##0.00 / #,##0.0000 / %)
        4. Автоподбор ширины колонок [10, 55]
        5. Закрепить первую строку (freeze_panes)
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    if df.empty:
        pd.DataFrame({"Информация": ["Нет данных для этой сессии"]}).to_excel(
            writer, sheet_name=sheet_name, index=False
        )
        return

    df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]

    # Стиль заголовка
    header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    header_font = Font(bold=True)
    header_align = Alignment(wrap_text=True, vertical="center", horizontal="center")

    for col_idx, col_name in enumerate(df.columns, 1):
        hdr_cell = ws.cell(1, col_idx)
        hdr_cell.fill  = header_fill
        hdr_cell.font  = header_font
        hdr_cell.alignment = header_align

        # Числовой формат для ячеек данных
        if pd.api.types.is_numeric_dtype(df[col_name]):
            fmt = _excel_num_format(col_name)
            for row_idx in range(2, len(df) + 2):
                cell = ws.cell(row_idx, col_idx)
                if cell.value is not None:
                    cell.number_format = fmt

    # Высота строки заголовка
    ws.row_dimensions[1].height = 32

    # Автоподбор ширины колонок
    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = max(
            len(str(col_name)),
            df[col_name].astype(str).str.len().max() if not df.empty else 0,
        )
        col_width = min(max(int(max_len) + 2, 10), 55)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Закрепить первую строку
    ws.freeze_panes = "A2"
